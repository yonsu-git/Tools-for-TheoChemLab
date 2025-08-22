import re
import os 
from pathlib import Path
from openpyxl import load_workbook


wb = load_workbook(r'c:\研究\ch2o_ice_cluster\ADS.xlsx')
ws = wb["H2CO"]

row_num_point = 2
row_num_sample_try = 2 
column = 1
col_num_point = 2
count = 0
zero = None
dir_path = r'c:\研究\ch2o_ice_cluster\log'

for path in Path(dir_path).iterdir():
    if path.is_file():
        sample = str(path.name)
        sample_try = sample.split("_")[2]
        print(sample_try) 
        
        sample_name = sample.split("_")[0]
        print(sample_name)
        
        if sample_name != ws.cell(row = 1, column = column).value:
            column += 1
            ws.cell(row = 1, column = column).value = sample_name

        if row_num_sample_try <= 11:
            ws[f"A{row_num_sample_try}"] = sample_try
            row_num_sample_try += 1

        print(path)
        
        with open(path, "r", encoding = "utf-8") as file:
            lines = file.readlines()
        for line in reversed(lines):
            if "Sum of electronic and zero-point Energies" in line:
                zero = float(line.split()[6])
                print(zero)
                break

    if zero is not None:
        print(zero)
        ws.cell(row = row_num_point, column = col_num_point).value = zero
        row_num_point += 1
        count += 1

        if count % 10 == 0:
            col_num_point += 1
            row_num_point = 2
    wb.save(r'c:\研究\ch2o_ice_cluster\ADS.xlsx')





